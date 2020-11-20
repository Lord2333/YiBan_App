import captcha_test
import CSDN_OCR
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from PIL import Image
import os
import requests

chrome_options = Options() # 实例化Option对象
chrome_options.add_argument('--headless') # 把Chrome浏览器设置为静默模式
#driver = webdriver.Chrome(options = chrome_options) # 设置引擎为Chrome，在后台默默运行
driver = webdriver.Chrome()#有窗口运行

Login_Url="http://www.yiban.cn/login?go=https://q.yiban.cn/app/index/appid/"#需要手动接上AppID

Vote_Data={
	"App_id": 780382,
	"Vote_id": 127102,
	"VoteOption_id[]": 1753786
}

def Xlsx_New():
	'''
	新建或覆盖当前目录下表格data.xlsx
	:return:
	'''
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = 'Students Data'
	sheet['A1'] = '提示：请严格按照以下格式填入数据！请把单元格格式转为文本！请勿删除本行！'
	rows = [['姓名', '账号', '密码'], ['张三', '13981012345', '220506'],
	        ['李四', '13981016789', '220506']]
	for i in rows:
		sheet.append(i)
	wb.save('data.xlsx')
	print("已在软件所在文件夹生成数据表格样本，请查看并添加数据！")

def Xlsx_Read(num):
	'''
	:param num:必须大于等于2
	:return: 依次返回昵称、账号、密码
	'''
	wb = openpyxl.load_workbook('./data.xlsx')#直接读取根目录下data.xlsx
	ws = wb.active
	rows = []
	for row in ws.iter_rows():
		rows.append(row)
	if num<2:
		return
	else:
		name = str(rows[num][0].value)  # 昵称
		account = str(rows[num][1].value)  # 账号
		password = str(rows[num][2].value)  # 密码
		return name,account,password

def Comparison(a,b):#暂停使用
	'''
	用来校验网页是否跳转，注意driver.current_url方法获取的链接末尾是否有/

	:param a: 链接1
	:param b: 链接2
	:return:返回1时说明两字符串相同，返回0时说明两字符串不同
	'''
	ib=0
	for ia in range(len(a)):
		if ord(a[ia:ia+1])-ord(b[ib:ib+1])==0:
			ib=ib+1
			if ib==len(b):
				return 1
			else:
				return 0

def Login(AppID,Acc_num):
	'''
	登录模块
	:param AppID:输入要进入的轻应用ID
	:param Acc_num: 账号序号，必须大于等于 2
	:return:
	'''
	driver.get(Login_Url+str(AppID))#拼接AppID并访问
	time.sleep(1)
	nicname,account,password=Xlsx_Read(Acc_num)
	Acc_Box=driver.find_element_by_id('account-txt')#找到账号输入框
	Pwd_Box=driver.find_element_by_id('password-txt')#找到密码输入框
	Acc_Box.send_keys(account)
	Pwd_Box.send_keys(password)
	Log_Buttom=driver.find_element_by_id('login-btn')#点击登录
	Log_Buttom.click()
	#如果点击登录后没有跳转，则引入验证码模块
	Web_Url = Login_Url + str(AppID)
	Web_Title=driver.title
	if Web_Title==u"易班 - 登录":#验证网页是否还在登录页
		Captcha_Buttom=driver.find_element_by_class_name('captcha')#找到验证码图片
		# 截图或验证码图片保存地址
		screenImg = "./screenImg.png"
		Dead_Num=1
		while True:
			Dead_Num=Dead_Num+1#防止死循环
			# 浏览器页面截屏
			driver.get_screenshot_as_file(screenImg)
			# 定位验证码位置及大小
			location = Captcha_Buttom.location
			size = Captcha_Buttom.size
			# 在网页截图的基础上重新裁切出来验证码
			left = location['x']
			top = location['y']
			right = location['x'] + size['width']
			bottom = location['y'] + size['height']
			#print(left+" "+right+" "+top+" "+bottom)
			# 从文件读取截图，截取验证码位置再次保存
			img = Image.open(screenImg).crop((left, top, right, bottom))
			img.save(screenImg)
			captcha_test.automation("screenImg.png")
			Ca_Text=CSDN_OCR.Recognise("transfered_image.png")
			Ca_Box=driver.find_element_by_id("login-captcha")
			Ca_Box.send_keys(Ca_Text)
			Log_Buttom.click()
			time.sleep(0.5)
			Web_Title=driver.title
			if Web_Title != u"易班 - 登录":
				print("登录成功")
				time.sleep(2)  # 等待两秒操作时间
				token = driver.get_cookie("yiban_user_token")
				token = token["yiban_user_token"]
				print(token)
				requests.post("https://q.yiban.cn/vote/insertBoxAjax", data=Vote_Data, cookies=token)
				print("投票成功！")
				time.sleep(1)
				driver.get("https://q.yiban.cn/logout")
				print("正在推出登录...")
				time.sleep(1)
				os.system("cls")
				driver.get(Login_Url + str(AppID))
				return 0 #中断函数
	elif Web_Title!=u"易班 - 登录":
		print("登录成功")
		time.sleep(2)  # 等待两秒操作时间
		token = driver.get_cookie()
		token =token["yiban_user_token"]
		print(token)
		requests.post("https://q.yiban.cn/vote/insertBoxAjax", data=Vote_Data, cookies=token)
		print("投票成功！")
		time.sleep(1)
		driver.get("https://q.yiban.cn/logout")
		print("正在推出登录...")
		time.sleep(1)
		os.system("cls")
		driver.get(Login_Url + str(AppID))
		return 0


if __name__ == '__main__':
	Xlsx_New()
	print("强制等待 五秒！")
	time.sleep(5)
	Send_Key=input("请保存表格后敲回车")
	#Login(764530,6)
	Stu_Num=int(input("请输入当前表格账号数量："))
	i=1
	while i < Stu_Num:
		i=i+1
		Login(780382,i)
		time.sleep(1)

	driver.close()