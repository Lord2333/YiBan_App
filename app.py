import time  # 导入时间库
import Copy #导入任务包
from lxml import etree
import openpyxl
import random
import requests

def Xlsx_New():
	'''
	新建或覆盖当前目录下表格data.xlsx
	:return:
	'''
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = 'Students Data'
	sheet['A1'] = '提示：请严格按照以下格式填入数据！请把单元格格式转为文本！请勿删除本行！'
	rows = [['姓名', '账号', '密码'], ['张三', '13981012345', '220506'],
	        ['李四', '13981016789', '220506']]
	for i in rows:
		sheet.append(i)
	wb.save('data.xlsx')
	print("已在软件所在文件夹生成数据表格样本，请查看并添加数据！")

def Xlsx_Read(num):
	'''
	:param num:必须大于等于2
	:return: 依次返回昵称、账号、密码
	'''
	wb = openpyxl.load_workbook('./data.xlsx')#直接读取根目录下data.xlsx
	ws = wb.active
	rows = []
	for row in ws.iter_rows():
		rows.append(row)
	if num<2:
		return
	else:
		name = str(rows[num][0].value)  # 昵称
		account = str(rows[num][1].value)  # 账号
		password = str(rows[num][2].value)  # 密码
		return name,account,password

global ddnum

def Res_Proxy(sheet_num):  # 定义一个循环的函数
	ddnum = 1
	# start_time = time.time()
	while ddnum < sheet_num:  # 当0<num=1时，一直执行以下代码
		# 设置请求头
		headers = {
			'User-Agent': "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.835.163 Safari/535.1"}
		# 依次遍历生成2-4
		for i in range(1, 999):
			url = "https://www.kuaidaili.com/free/inha/" + str(i) + "/"  # 爬取的免费ip
			response = requests.get(url, headers=headers).text  # 获得网页文本数据
			response_xpath = etree.HTML(response)  # 转换为xpath可用结构
			ips = response_xpath.xpath('//*[@id="list"]/table/tbody/tr/td[1]/text()')  # ip的信息
			dks = response_xpath.xpath('//*[@id="list"]/table/tbody/tr/td[2]/text()')  # 端口的信息
			https = response_xpath.xpath('//*[@id="list"]/table/tbody/tr[2]/td[4]/text()')  # http信息
			for ip, dk, http in zip(ips, dks, https):
				ddnum = ddnum + 1
				nicname, account, password = Xlsx_Read(ddnum)
				proxy = "http://" + ip + ":" + dk  # 拼接ip
				# print(proxy)
				proxies = {"http": proxy}
				Copy.login(nicname, account, password,)
				Ra_Num = random.randrange(10, 60, 5)
				print("暂停{:}秒".format(Ra_Num))  # 后续把这里改成随机数
				time.sleep(Ra_Num)

if __name__=="__main__":
	stime = time.time()
	Xlsx_New()
	sheet_num=int(input("请输入表格账号数量："))
	Res_Proxy(sheet_num)
	etime = time.time()
	atime = etime - stime
	min = atime / 60
	sec = atime % 60
	print("本次任务耗时%d分%d秒！\n即将关闭程序" % (min, sec))